import io
import json
from collections import defaultdict
from typing import Any, Dict, List, Tuple

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from rest_framework import mixins, permissions, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination

from .constants import TEMPLATE_FIELDS
from .models import FormSubmission, FormTemplate
from .serializers import FormSubmissionSerializer, FormTemplateSerializer


class StandardPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"


class FormTemplateViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    queryset = FormTemplate.objects.all()
    serializer_class = FormTemplateSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = "template_type"


class FormSubmissionViewSet(viewsets.ModelViewSet):
    serializer_class = FormSubmissionSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardPagination

    def get_permissions(self):
        return [permissions.AllowAny()]
        # Cho phép gửi form hoặc xem chi tiết submission không cần đăng nhập; các thao tác khác yêu cầu auth.
        if self.action in {"create", "retrieve"}:
            return [permissions.AllowAny()]
        return super().get_permissions()

    def get_queryset(self):
        base = FormSubmission.objects.select_related("template")

        template_type = self.request.query_params.get("template_type")
        if template_type:
            base = base.filter(template_type=template_type)
        status_param = self.request.query_params.get("status")
        if status_param:
            base = base.filter(status=status_param)
        return base

    def _format_section(self, value: Any, field_labels: Dict[str, str]) -> str:
        """Human-readable string for a section (list/dict/scalar)."""
        placeholder = "Không điền"
        if isinstance(value, list):
            formatted_rows = []
            for idx, row in enumerate(value, start=1):
                if not isinstance(row, dict):
                    continue
                row_parts = []
                for key, label in field_labels.items():
                    cell = row.get(key)
                    display = placeholder if cell in (None, "") else cell
                    row_parts.append(f"{label}: {display}")
                if row_parts:
                    formatted_rows.append(f"#{idx} " + "; ".join(row_parts))
            return "\n".join(formatted_rows)

        if isinstance(value, dict):
            row_parts = []
            for key, label in field_labels.items():
                cell = value.get(key)
                display = placeholder if cell in (None, "") else cell
                row_parts.append(f"{label}: {display}")
            return "; ".join(row_parts)

        return placeholder if value in (None, "") else str(value)

    def _format_submission_data(self, submission: FormSubmission) -> str:
        mapping = TEMPLATE_FIELDS.get(submission.template_type)
        data = submission.data or {}
        if not mapping:
            return json.dumps(data, ensure_ascii=False)

        sections_output = []
        for section_key, field_labels in mapping.items():
            if section_key not in data:
                continue
            formatted = self._format_section(
                data.get(section_key), field_labels)
            if formatted:
                sections_output.append(f"[{section_key}] {formatted}")

        remaining = {k: v for k, v in data.items() if k not in mapping}
        if remaining:
            sections_output.append(
                "[khac] " + json.dumps(remaining, ensure_ascii=False))

        return "\n".join(sections_output)

    def _template_extra_headers(self, template_type: str) -> List[Tuple[str, str]]:
        """Generate extra column keys and headers for a template based on field labels."""
        mapping = TEMPLATE_FIELDS.get(template_type)
        if not mapping:
            return []

        extra: List[Tuple[str, str]] = []
        for section_key, field_labels in mapping.items():
            for field_key, label in field_labels.items():
                col_key = f"{section_key}.{field_key}"
                header = f"{section_key.upper()} - {label}"
                extra.append((col_key, header))
        return extra

    def _format_field_value(self, section_value: Any, field_key: str) -> str:
        """Extract and stringify a single field from a section value."""
        placeholder = "Không điền"
        if isinstance(section_value, list):
            parts = []
            for idx, row in enumerate(section_value, start=1):
                if isinstance(row, dict):
                    val = row.get(field_key)
                    display = placeholder if val in (None, "") else val
                    parts.append(f"#{idx}: {display}")
            return " | ".join(parts)
        if isinstance(section_value, dict):
            val = section_value.get(field_key)
            return placeholder if val in (None, "") else str(val)
        return placeholder if section_value in (None, "") else str(section_value)

    def _flatten_submission_columns(self, submission: FormSubmission, columns: List[Tuple[str, str]]) -> Dict[str, str]:
        data = submission.data or {}
        flat: Dict[str, str] = {}
        for col_key, _ in columns:
            if "." not in col_key:
                flat[col_key] = ""
                continue
            section_key, field_key = col_key.split(".", 1)
            flat[col_key] = self._format_field_value(
                data.get(section_key), field_key)
        return flat

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        submissions = self.get_queryset().order_by("-created_at")

        # Group submissions by template type so each template has its own sheet.
        grouped = defaultdict(list)
        for submission in submissions:
            grouped[submission.template_type].append(submission)

        wb = Workbook()
        default_sheet = wb.active
        default_sheet.title = "__placeholder__"

        base_headers = [
            "STT",
            "Loại Mẫu Form",
            "Họ Tên",
            "Năm Sinh",
            "Số Điện Thoại",
            "Thôn",
            "Xã",
            "Tỉnh",
            "Tình Trạng",
            "Ngày Tạo",
        ]

        for template_type, items in grouped.items():
            extra_headers = self._template_extra_headers(template_type)
            headers = base_headers + [h for _, h in extra_headers]
            ws = wb.create_sheet(title=template_type[:31])
            ws.append(headers)

            for idx, submission in enumerate(items, start=1):
                flat_columns = self._flatten_submission_columns(
                    submission, extra_headers)
                ws.append(
                    [
                        idx,
                        submission.template_type,
                        submission.ho_ten,
                        submission.nam_sinh,
                        submission.so_dien_thoai,
                        submission.thon,
                        submission.xa,
                        submission.tinh,
                        submission.status,
                        submission.created_at.isoformat(),
                        *[flat_columns.get(key, "") for key, _ in extra_headers],
                    ]
                )

            for idx, header in enumerate(headers, start=1):
                column_letter = get_column_letter(idx)
                ws.column_dimensions[column_letter].width = max(
                    min(len(str(header)) + 2, 40), 18
                )

            info_start = len(base_headers) + 1
            if len(headers) >= info_start:
                for row in ws.iter_rows(min_row=2, min_col=info_start, max_col=len(headers)):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        if default_sheet in wb.worksheets and default_sheet.title == "__placeholder__":
            wb.remove(default_sheet)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"form_submissions_{timezone.now():%Y%m%d%H%M%S}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
